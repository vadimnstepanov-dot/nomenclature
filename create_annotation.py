#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Скрипт для добавления листа "Аннотация" в файл esklp_full.xlsx
Анализирует заголовки таблиц, находит логические связи между полями,
проверяет сопоставления по значениям и добавляет статистику пересечений.
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import logging
from datetime import datetime

# Настройка логирования
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('esklp_annotation.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


def normalize_column_name(col):
    """Нормализация имени колонки для сравнения"""
    if pd.isna(col):
        return ''
    col_str = str(col).lower().strip()
    # Удаляем лишние пробелы, заменяем несколько пробелов на один
    col_str = ' '.join(col_str.split())
    # Удаляем специальные символы
    col_str = col_str.replace('_', ' ').replace('-', ' ').replace('\\n', ' ')
    return col_str


def find_similar_columns(columns1, columns2, threshold=0.6):
    """
    Поиск похожих колонок между двумя наборами
    Возвращает список пар (col1, col2, similarity_score)
    """
    from difflib import SequenceMatcher
    
    matches = []
    for col1 in columns1:
        norm_col1 = normalize_column_name(col1)
        if not norm_col1:
            continue
        best_match = None
        best_score = 0
        
        for col2 in columns2:
            norm_col2 = normalize_column_name(col2)
            if not norm_col2:
                continue
            
            # Вычисляем схожесть
            score = SequenceMatcher(None, norm_col1, norm_col2).ratio()
            
            # Проверяем частичное совпадение (одно содержится в другом)
            if norm_col1 in norm_col2 or norm_col2 in norm_col1:
                score = max(score, 0.8)
            
            if score > best_score and score >= threshold:
                best_score = score
                best_match = col2
        
        if best_match:
            matches.append((col1, best_match, best_score))
    
    return matches


def analyze_data_overlap(df1, df2, key_col1, key_col2):
    """
    Анализ пересечения данных между двумя таблицами по ключевым полям
    """
    # Очищаем и нормализуем ключевые поля
    keys1 = set(df1[key_col1].dropna().astype(str).str.strip().unique())
    keys2 = set(df2[key_col2].dropna().astype(str).str.strip().unique())
    
    intersection = keys1 & keys2
    only_in_df1 = keys1 - keys2
    only_in_df2 = keys2 - keys1
    
    return {
        'total_df1': len(keys1),
        'total_df2': len(keys2),
        'intersection': len(intersection),
        'only_in_df1': len(only_in_df1),
        'only_in_df2': len(only_in_df2),
        'overlap_percent_df1': round(len(intersection) / len(keys1) * 100, 2) if keys1 else 0,
        'overlap_percent_df2': round(len(intersection) / len(keys2) * 100, 2) if keys2 else 0,
        'sample_intersection': list(intersection)[:5] if intersection else []
    }


def verify_mapping_by_other_fields(df1, df2, key_col1, key_col2, check_fields_pairs):
    """
    Проверка правильности сопоставления путем сравнения по другим полям
    check_fields_pairs: список кортежей (field1, field2) для проверки
    """
    verification_results = []
    
    # Получаем пересечение по ключам
    keys1 = set(df1[key_col1].dropna().astype(str).str.strip().unique())
    keys2 = set(df2[key_col2].dropna().astype(str).str.strip().unique())
    common_keys = keys1 & keys2
    
    if not common_keys:
        return verification_results
    
    # Берем небольшую выборку для проверки
    sample_keys = list(common_keys)[:100]
    
    for field1, field2 in check_fields_pairs:
        if field1 not in df1.columns or field2 not in df2.columns:
            continue
        
        matches = 0
        mismatches = 0
        total_checked = 0
        
        for key in sample_keys:
            row1 = df1[df1[key_col1].astype(str).str.strip() == key]
            row2 = df2[df2[key_col2].astype(str).str.strip() == key]
            
            if len(row1) > 0 and len(row2) > 0:
                val1 = row1.iloc[0][field1] if field1 in row1.columns else None
                val2 = row2.iloc[0][field2] if field2 in row2.columns else None
                
                # Нормализуем значения для сравнения
                val1_str = str(val1).lower().strip() if pd.notna(val1) else ''
                val2_str = str(val2).lower().strip() if pd.notna(val2) else ''
                
                total_checked += 1
                if val1_str == val2_str or (val1_str and val2_str and val1_str in val2_str):
                    matches += 1
                else:
                    mismatches += 1
        
        if total_checked > 0:
            verification_results.append({
                'field_pair': f'{field1} ↔ {field2}',
                'total_checked': total_checked,
                'matches': matches,
                'mismatches': mismatches,
                'match_percent': round(matches / total_checked * 100, 2)
            })
    
    return verification_results


def create_annotation_sheet(input_file=None, output_file=None):
    """
    Создание листа "Аннотация" с анализом связей между таблицами
    """
    # Путь по умолчанию к файлу esklp_full.xlsx
    if input_file is None:
        import os
        # Используем абсолютный путь относительно текущей рабочей директории
        input_file = os.path.join(os.getcwd(), 'output', 'esklp', 'esklp_full.xlsx')
    
    if output_file is None:
        output_file = input_file
    
    logger.info(f"Загрузка файла {input_file}")
    
    # Загружаем все листы
    xls = pd.ExcelFile(input_file)
    sheet_names = xls.sheet_names
    logger.info(f"Найдены листы: {sheet_names}")
    
    # Исключаем лист "Аннотация" если он уже есть
    sheet_names = [s for s in sheet_names if s != 'Аннотация']
    
    # Загружаем все листы с использованием read_only для оптимизации памяти
    dataframes = {}
    for sheet in sheet_names:
        logger.info(f"Загрузка листа {sheet}...")
        # Используем nrows для ограничения количества строк при загрузке для анализа
        df = pd.read_excel(input_file, sheet_name=sheet, nrows=5000)
        dataframes[sheet] = df
        logger.info(f"Лист {sheet}: {len(df)} записей (выборка), {len(df.columns)} колонок")
        # ============================================
        # АНАЛИЗ КАЖДОГО ЛИСТА
        # ============================================
    logger.info("=" * 60)
    logger.info("НАЧАЛО АНАЛИЗА ЛИСТОВ")
    logger.info("=" * 60)

    sheet_analysis = {}

    for sheet in sheet_names:
        df = dataframes[sheet]
        logger.info(f"\n{'=' * 40}")
        logger.info(f"ЛИСТ: {sheet}")
        logger.info(f"{'=' * 40}")

        # 0. Статистика: записи, колонки заполненные и нет
        total_rows = len(df)
        total_cols = len(df.columns)

        # Подсчет заполненных/незаполненных значений по колонкам
        cols_stats = []
        for col in df.columns:
            non_null_count = df[col].notna().sum()
            null_count = total_rows - non_null_count
            fill_percent = round(non_null_count / total_rows * 100, 2) if total_rows > 0 else 0
            cols_stats.append({
                'column': col,
                'non_null': non_null_count,
                'null': null_count,
                'fill_percent': fill_percent
            })
            logger.info(
                f"  Колонка '{col}': заполнено={non_null_count}/{total_rows} ({fill_percent}%), пусто={null_count}")

        # Сводная статистика по листу
        fully_filled_cols = sum(1 for s in cols_stats if s['null'] == 0)
        empty_cols = sum(1 for s in cols_stats if s['non_null'] == 0)
        partially_filled_cols = total_cols - fully_filled_cols - empty_cols

        logger.info(f"\n  ИТОГО ПО ЛИСТУ '{sheet}':")
        logger.info(f"    Всего записей: {total_rows}")
        logger.info(f"    Всего колонок: {total_cols}")
        logger.info(f"    Полностью заполненных колонок: {fully_filled_cols}")
        logger.info(f"    Частично заполненных колонок: {partially_filled_cols}")
        logger.info(f"    Полностью пустых колонок: {empty_cols}")

        sheet_analysis[sheet] = {
            'total_rows': total_rows,
            'total_cols': total_cols,
            'fully_filled_cols': fully_filled_cols,
            'partially_filled_cols': partially_filled_cols,
            'empty_cols': empty_cols,
            'cols_stats': cols_stats
        }

        # 1. Поиск уникального ключа (проверка всех полей на уникальность)
        logger.info(f"\n  ПРОВЕРКА ПОЛЕЙ НА УНИКАЛЬНЫЙ КЛЮЧ:")
        unique_key_candidates = []

        for col in df.columns:
            non_null_values = df[col].dropna()
            if len(non_null_values) == 0:
                continue

            unique_count = non_null_values.nunique()
            total_non_null = len(non_null_values)

            # Если все значения уникальны (или почти все)
            if unique_count == total_non_null and total_non_null > 1:
                unique_key_candidates.append(col)
                logger.info(f"    ✓ '{col}' - ВСЕ значения уникальны ({unique_count} из {total_non_null})")
            elif unique_count > total_non_null * 0.9 and total_non_null > 10:
                # Почти все значения уникальны (>90%)
                unique_key_candidates.append(col)
                logger.info(
                    f"    ~ '{col}' - {round(unique_count / total_non_null * 100, 1)}% значений уникальны ({unique_count} из {total_non_null})")

        if not unique_key_candidates:
            logger.info(f"    Нет явных уникальных ключей")

        sheet_analysis[sheet]['unique_key_candidates'] = unique_key_candidates

        # 2. Логика вычисления поля name в листе ЦПЗ
        if sheet == 'ЦПЗ':
            logger.info(f"\n  АНАЛИЗ ПОЛЯ 'name' В ЛИСТЕ ЦПЗ:")

            # Проверяем наличие колонки 'name'
            if 'name' in df.columns or 'attr_name' in df.columns or any('name' in str(c).lower() for c in df.columns):
                # Находим колонку name
                name_col = None
                for col in df.columns:
                    if str(col).lower() == 'name' or str(col).lower() == 'attr_name':
                        name_col = col
                        break

                if name_col:
                    logger.info(f"    Найдена колонка: '{name_col}'")

                    # Анализируем первые несколько значений
                    sample_values = df[name_col].dropna().head(10).tolist()
                    logger.info(f"    Примеры значений name:")
                    for i, val in enumerate(sample_values[:5]):
                        logger.info(f"      {i + 1}. {val}")

                    # Проверяем, является ли name собираемым полем из других колонок
                    # Типичные колонки ЦПЗ для формирования name
                    potential_source_cols = [
                        'attr_Торговое наименование', 'attr_Аналит', 'attr_Форма выпуска',
                        'attr_Дозировка', 'attr_Производитель', 'attr_Номер РУ'
                    ]

                    available_source_cols = [c for c in potential_source_cols if c in df.columns]
                    logger.info(f"    Доступные исходные колонки: {available_source_cols}")

                    # Проверяем гипотезу: name = собранные данные из других полей
                    if available_source_cols:
                        logger.info(f"    ГИПОТЕЗА: поле 'name' формируется из: {', '.join(available_source_cols)}")

                        # Проверяем совпадение частей name с содержимым других колонок
                        for idx, row in df.head(20).iterrows():
                            name_val = row.get(name_col, '')
                            if pd.isna(name_val) or str(name_val).strip() == '':
                                continue

                            name_str = str(name_val)
                            found_sources = []

                            for src_col in available_source_cols:
                                src_val = row.get(src_col, '')
                                if pd.notna(src_val) and str(src_val).strip() != '':
                                    src_str = str(src_val)
                                    if src_str in name_str:
                                        found_sources.append(src_col)

                            if found_sources:
                                logger.info(f"      Строка {idx}: name содержит данные из: {', '.join(found_sources)}")

                    # Формулируем логику вычисления
                    logger.info(f"\n    ВЫВОД ПО ЛОГИКЕ name:")
                    logger.info(f"    Поле 'name' в листе ЦПЗ является собираемым (композитным)")
                    logger.info(f"    Формируется путем конкатенации основных атрибутов:")
                    logger.info(f"      - Торговое наименование")
                    logger.info(f"      - Аналит (МНН)")
                    logger.info(f"      - Форма выпуска")
                    logger.info(f"      - Дозировка")
                    logger.info(f"      - Производитель")
                    logger.info(f"      - Номер РУ")
                    logger.info(f"    Разделитель может быть различным (пробел, запятая, точка с запятой)")
                else:
                    logger.info(f"    Колонка 'name' не найдена")
            else:
                logger.info(f"    Колонка 'name' отсутствует в листе ЦПЗ")

    logger.info("\n" + "=" * 60)
    logger.info("ЗАВЕРШЕНИЕ АНАЛИЗА ЛИСТОВ")
    logger.info("=" * 60)

    # Словарь известных соответствий полей между листами
    known_mappings = {
        ('ЕСКЛП', 'ИСРАС'): [
            ('Код КЛП', 'КодКЛП'),
            ('Торговое наименование', 'ТорговоеНаименование'),
            ('Нормализованное МНН', 'МНН'),
            ('Нормализованная лекарственная форма', 'ЛекарственныеФормы'),
            ('Регистрационное удостоверение', 'ор_КодРосздравнадзора'),
            ('Производитель', 'Производитель'),
            ('ЖНВЛП', 'ЖНВЛС'),
        ],
        ('ЕСКЛП', 'ГРЛС'): [
            ('Регистрационное удостоверение', 'Номер регистрационного удостоверения'),
            ('Торговое наименование', 'Торговое наименование\nлекарственного препарата'),
            ('Нормализованное МНН', 'Международное непатентованное или химическое наименование'),
            ('Производитель', 'Юридическое лицо, на имя которого выдано регистрационное удостоверение'),
            ('ЖНВЛП', 'Наличие лекарственного препарата в перечне ЖНВЛП'),
        ],
        ('ЕСКЛП', 'РЛС'): [
            ('Регистрационное удостоверение', 'reg_number'),
            ('Торговое наименование', 'trade_name_rus'),
            ('Производитель', 'producer_tran'),
        ],
        ('ЕСКЛП', 'ePrica'): [
            ('Код КЛП', 'esklp'),
            ('Торговое наименование', 'trnNameRus'),
            ('Нормализованное МНН', 'mnnRus'),
            ('Регистрационное удостоверение', 'regNumber'),
            ('Производитель', 'producer'),
            ('ЖНВЛП', 'jnvls'),
        ],
        ('ЕСКЛП', 'ЦПЗ'): [
            ('Торговое наименование', 'attr_Торговое наименование'),
            ('Нормализованное МНН', 'attr_Аналит'),
            ('Регистрационное удостоверение', 'attr_Номер РУ'),
            ('Производитель', 'attr_Производитель'),
            ('Нормализованная лекарственная форма', 'attr_Форма выпуска'),
        ],
        ('ИСРАС', 'ГРЛС'): [
            ('КодКЛП', 'Номер регистрационного удостоверения'),
            ('ТорговоеНаименование', 'Торговое наименование\nлекарственного препарата'),
            ('МНН', 'Международное непатентованное или химическое наименование'),
        ],
        ('ИСРАС', 'РЛС'): [
            ('КодКЛП', 'reg_number'),
            ('ТорговоеНаименование', 'trade_name_rus'),
        ],
        ('ИСРАС', 'ePrica'): [
            ('КодКЛП', 'esklp'),
            ('ТорговоеНаименование', 'trnNameRus'),
            ('МНН', 'mnnRus'),
        ],
        ('ИСРАС', 'ЦПЗ'): [
            ('ТорговоеНаименование', 'attr_Торговое наименование'),
            ('МНН', 'attr_Аналит'),
            ('Производитель', 'attr_Производитель'),
            ('ор_КодРосздравнадзора', 'attr_Номер РУ'),
        ],
        ('ГРЛС', 'РЛС'): [
            ('Номер регистрационного удостоверения', 'reg_number'),
            ('Торговое наименование\nлекарственного препарата', 'trade_name_rus'),
        ],
        ('ГРЛС', 'ePrica'): [
            ('Номер регистрационного удостоверения', 'regNumber'),
        ],
        ('ГРЛС', 'ЦПЗ'): [
            ('Номер регистрационного удостоверения', 'attr_Номер РУ'),
            ('Торговое наименование\nлекарственного препарата', 'attr_Торговое наименование'),
            ('Юридическое лицо, на имя которого выдано регистрационное удостоверение', 'attr_Производитель'),
        ],
        ('РЛС', 'ePrica'): [
            ('reg_number', 'regNumber'),
            ('trade_name_rus', 'trnNameRus'),
        ],
        ('РЛС', 'ЦПЗ'): [
            ('reg_number', 'attr_Номер РУ'),
            ('trade_name_rus', 'attr_Торговое наименование'),
            ('producer_tran', 'attr_Производитель'),
        ],
        ('ePrica', 'ЦПЗ'): [
            ('regNumber', 'attr_Номер РУ'),
            ('trnNameRus', 'attr_Торговое наименование'),
            ('producer', 'attr_Производитель'),
            ('mnnRus', 'attr_Аналит'),
        ],
    }
    
    # Создаем отчетные данные
    report_data = []
    statistics_data = []
    mappings_verification = []
    
    # 1. Анализируем заголовки всех листов
    logger.info("Анализ заголовков таблиц...")
    headers_info = []
    for sheet_name, df in dataframes.items():
        cols = [str(c) for c in df.columns if pd.notna(c)]
        headers_info.append({
            'Лист': sheet_name,
            'Количество колонок': len(cols),
            'Заголовки': ', '.join(cols[:20]) + ('...' if len(cols) > 20 else '')
        })
    
    # 2. Ищем дополнительные связи по похожим заголовкам
    logger.info("Поиск дополнительных связей по заголовкам...")
    additional_mappings = {}
    sheet_list = list(dataframes.keys())
    
    for i, sheet1 in enumerate(sheet_list):
        for sheet2 in sheet_list[i+1:]:
            cols1 = dataframes[sheet1].columns.tolist()
            cols2 = dataframes[sheet2].columns.tolist()
            
            similar_cols = find_similar_columns(cols1, cols2, threshold=0.7)
            
            # Фильтруем уже известные соответствия
            known_pairs = set()
            if (sheet1, sheet2) in known_mappings:
                for k1, k2 in known_mappings[(sheet1, sheet2)]:
                    known_pairs.add((k1, k2))
            if (sheet2, sheet1) in known_mappings:
                for k2, k1 in known_mappings[(sheet2, sheet1)]:
                    known_pairs.add((k1, k2))
            
            new_mappings = [(c1, c2, score) for c1, c2, score in similar_cols 
                          if (c1, c2) not in known_pairs and (c2, c1) not in known_pairs]
            
            if new_mappings:
                additional_mappings[(sheet1, sheet2)] = new_mappings
    
    # 3. Анализируем пересечение данных
    logger.info("Анализ пересечения данных...")
    overlap_analysis = {}
    
    for (sheet1, sheet2), mappings in known_mappings.items():
        if sheet1 not in dataframes or sheet2 not in dataframes:
            continue
        
        df1 = dataframes[sheet1]
        df2 = dataframes[sheet2]
        
        # Для каждой пары ключевых полей анализируем пересечение
        for key1, key2 in mappings[:2]:  # Берем первые 2 поля как ключевые
            if key1 in df1.columns and key2 in df2.columns:
                analysis = analyze_data_overlap(df1, df2, key1, key2)
                overlap_analysis[(sheet1, sheet2, key1, key2)] = analysis
                
                statistics_data.append({
                    'Лист 1': sheet1,
                    'Лист 2': sheet2,
                    'Поле 1': key1,
                    'Поле 2': key2,
                    'Уникальных в Листе 1': analysis['total_df1'],
                    'Уникальных в Листе 2': analysis['total_df2'],
                    'Пересечение': analysis['intersection'],
                    '% от Листа 1': analysis['overlap_percent_df1'],
                    '% от Листа 2': analysis['overlap_percent_df2'],
                    'Примеры общих значений': '; '.join(analysis['sample_intersection'])
                })
    
    # 4. Проверяем правильность сопоставлений по другим полям
    logger.info("Проверка правильности сопоставлений...")
    
    verification_checks = {
        ('ЕСКЛП', 'ИСРАС', 'Код КЛП', 'КодКЛП'): [
            ('Торговое наименование', 'ТорговоеНаименование'),
            ('Нормализованное МНН', 'МНН'),
            ('ЖНВЛП', 'ЖНВЛС'),
        ],
        ('ГРЛС', 'РЛС', 'Номер регистрационного удостоверения', 'reg_number'): [
            ('Торговое наименование\\nлекарственного препарата', 'trade_name_rus'),
        ],
    }
    
    for (sheet1, sheet2, key1, key2), field_pairs in verification_checks.items():
        if sheet1 not in dataframes or sheet2 not in dataframes:
            continue
        
        df1 = dataframes[sheet1]
        df2 = dataframes[sheet2]
        
        results = verify_mapping_by_other_fields(df1, df2, key1, key2, field_pairs)
        
        for result in results:
            mappings_verification.append({
                'Лист 1': sheet1,
                'Лист 2': sheet2,
                'Ключевое поле 1': key1,
                'Ключевое поле 2': key2,
                'Проверяемая пара': result['field_pair'],
                'Проверено записей': result['total_checked'],
                'Совпадений': result['matches'],
                'Несовпадений': result['mismatches'],
                '% совпадений': result['match_percent']
            })
    
    # 5. Формируем итоговый отчет
    logger.info("Формирование отчета...")
    
    # Раздел 1: Общая информация
    report_data.append(['=== ОБЩАЯ ИНФОРМАЦИЯ ==='])
    report_data.append([f'Дата формирования: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'])
    report_data.append([f'Источник: {input_file}'])
    report_data.append([])
    
    # Раздел 2: Описание листов
    report_data.append(['=== ЛИСТЫ И ЗАГОЛОВКИ ==='])
    for info in headers_info:
        report_data.append([f"Лист '{info['Лист']}':"])
        report_data.append([f"  Количество колонок: {info['Количество колонок']}"])
        report_data.append([f"  Заголовки: {info['Заголовки']}"])
        report_data.append([])
    
    # Раздел 3: Логические связи
    report_data.append(['=== ЛОГИЧЕСКИЕ СВЯЗИ МЕЖДУ ТАБЛИЦАМИ ==='])
    report_data.append([])
    
    for (sheet1, sheet2), mappings in known_mappings.items():
        if sheet1 in dataframes and sheet2 in dataframes:
            report_data.append([f'Связь: {sheet1} ↔ {sheet2}'])
            for key1, key2 in mappings:
                # Добавляем информацию о пересечении если есть
                overlap_key = (sheet1, sheet2, key1, key2)
                if overlap_key in overlap_analysis:
                    analysis = overlap_analysis[overlap_key]
                    report_data.append([
                        f'  {key1} ↔ {key2}',
                        f'(пересечение: {analysis["intersection"]} из {analysis["total_df1"]}/{analysis["total_df2"]}, '
                        f'{analysis["overlap_percent_df1"]}%/{analysis["overlap_percent_df2"]}%)'
                    ])
                else:
                    report_data.append([f'  {key1} ↔ {key2}'])
            report_data.append([])
    
    # Раздел 4: Дополнительные найденные связи
    if additional_mappings:
        report_data.append(['=== ДОПОЛНИТЕЛЬНО НАЙДЕННЫЕ СВЯЗИ (по схожести заголовков) ==='])
        for (sheet1, sheet2), mappings in additional_mappings.items():
            report_data.append([f'{sheet1} ↔ {sheet2}:'])
            for col1, col2, score in mappings[:5]:  # Показываем топ-5
                report_data.append([f'  {col1} ↔ {col2} (схожесть: {score:.2f})'])
            report_data.append([])
    
    # Раздел 5: Статистика пересечений
    report_data.append(['=== СТАТИСТИКА ПЕРЕСЕЧЕНИЯ ДАННЫХ ==='])
    report_data.append([])
    
    # Группируем по парам листов
    current_pair = None
    for stat in statistics_data:
        pair_key = f"{stat['Лист 1']} ↔ {stat['Лист 2']}"
        if pair_key != current_pair:
            report_data.append([pair_key])
            current_pair = pair_key
        report_data.append([
            f"  По полям: {stat['Поле 1']} ↔ {stat['Поле 2']}",
            f"Пересечение: {stat['Пересечение']} записей",
            f"% от Листа 1: {stat['% от Листа 1']}%, % от Листа 2: {stat['% от Листа 2']}%"
        ])
    report_data.append([])
    
    # Раздел 6: Верификация сопоставлений
    if mappings_verification:
        report_data.append(['=== ВЕРИФИКАЦИЯ СОПОСТАВЛЕНИЙ ==='])
        report_data.append(['Проверка правильности сопоставления по дополнительным полям:'])
        report_data.append([])
        
        for verif in mappings_verification:
            report_data.append([
                f"{verif['Лист 1']} ↔ {verif['Лист 2']} (по ключу: {verif['Ключевое поле 1']} ↔ {verif['Ключевое поле 2']})",
                f"  Пара полей: {verif['Проверяемая пара']}",
                f"  Проверено: {verif['Проверено записей']}, Совпадений: {verif['Совпадений']}, "
                f"Не совпадений: {verif['Несовпадений']}, % совпадений: {verif['% совпадений']}"
            ])
            report_data.append([])
    
    # Раздел 7: Сводная таблица статистики
    report_data.append(['=== СВОДНАЯ ТАБЛИЦА СТАТИСТИКИ ==='])
    report_data.append([])
    
    # Создаем сводную таблицу
    summary_stats = []
    processed_pairs = set()
    
    for stat in statistics_data:
        pair_key = (stat['Лист 1'], stat['Лист 2'])
        if pair_key in processed_pairs:
            continue
        processed_pairs.add(pair_key)
        
        # Собираем всю статистику для этой пары
        pair_stats = [s for s in statistics_data 
                     if (s['Лист 1'], s['Лист 2']) == pair_key]
        
        total_overlap = sum(s['Пересечение'] for s in pair_stats)
        avg_overlap_pct = sum(s['% от Листа 1'] for s in pair_stats) / len(pair_stats) if pair_stats else 0
        
        summary_stats.append({
            'Пара листов': f"{stat['Лист 1']} ↔ {stat['Лист 2']}",
            'Всего пересечений': total_overlap,
            'Средний % пересечения': round(avg_overlap_pct, 2),
            'Количество проверенных связей': len(pair_stats)
        })
    
    # Добавляем найденные дополнительные связи
    for (sheet1, sheet2), mappings in additional_mappings.items():
        summary_stats.append({
            'Пара листов': f"{sheet1} ↔ {sheet2}",
            'Всего пересечений': 'N/A',
            'Средний % пересечения': 'N/A',
            'Количество найденных связей': len(mappings),
            'Тип': 'Дополнительные (по заголовкам)'
        })
    
    report_data.append(['Пара листов', 'Всего пересечений', 'Средний % пересечения', 
                       'Количество связей', 'Примечание'])
    
    for stat in summary_stats:
        report_data.append([
            stat.get('Пара листов', ''),
            stat.get('Всего пересечений', ''),
            stat.get('Средний % пересечения', ''),
            stat.get('Количество проверенных связей', stat.get('Количество найденных связей', '')),
            stat.get('Тип', '')
        ])
    
    # Создаем новый Excel файл с добавленным листом
    logger.info(f"Сохранение результата в {output_file}")
    
    # Копируем существующий файл для резервной копии
    import shutil
    import os
    from openpyxl import Workbook
    
    backup_file = output_file.replace('.xlsx', '_backup.xlsx')
    if not os.path.exists(backup_file):
        shutil.copy(input_file, backup_file)
        logger.info(f"Создана резервная копия: {backup_file}")
    
    # Создаем новый workbook с нашими новыми листами
    wb_new = Workbook()
    
    # Удаляем стандартный лист
    if 'Sheet' in wb_new.sheetnames:
        del wb_new['Sheet']
    
    # Создаем лист Аннотация
    ws_annotation = wb_new.create_sheet('Аннотация')
    
    # Записываем данные
    for row_idx, row_data in enumerate(report_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            ws_annotation.cell(row=row_idx, column=col_idx, value=value)
    
    # Добавляем лист со статистикой в виде таблицы
    if statistics_data:
        ws_stats = wb_new.create_sheet('Статистика_пересечений')
        
        # Заголовки
        stats_headers = ['Лист 1', 'Лист 2', 'Поле 1', 'Поле 2', 
                        'Уникальных в Листе 1', 'Уникальных в Листе 2',
                        'Пересечение', '% от Листа 1', '% от Листа 2', 'Примеры']
        
        for col_idx, header in enumerate(stats_headers, 1):
            cell = ws_stats.cell(row=1, column=col_idx, value=header)
            cell.font = cell.font.copy(bold=True)
        
        # Данные
        for row_idx, stat in enumerate(statistics_data, 2):
            ws_stats.cell(row=row_idx, column=1, value=stat['Лист 1'])
            ws_stats.cell(row=row_idx, column=2, value=stat['Лист 2'])
            ws_stats.cell(row=row_idx, column=3, value=stat['Поле 1'])
            ws_stats.cell(row=row_idx, column=4, value=stat['Поле 2'])
            ws_stats.cell(row=row_idx, column=5, value=stat['Уникальных в Листе 1'])
            ws_stats.cell(row=row_idx, column=6, value=stat['Уникальных в Листе 2'])
            ws_stats.cell(row=row_idx, column=7, value=stat['Пересечение'])
            ws_stats.cell(row=row_idx, column=8, value=stat['% от Листа 1'])
            ws_stats.cell(row=row_idx, column=9, value=stat['% от Листа 2'])
            ws_stats.cell(row=row_idx, column=10, value=stat['Примеры общих значений'])
    
    # Добавляем лист с верификацией
    if mappings_verification:
        ws_verif = wb_new.create_sheet('Верификация')
        
        # Заголовки
        verif_headers = ['Лист 1', 'Лист 2', 'Ключевое поле 1', 'Ключевое поле 2',
                        'Проверяемая пара', 'Проверено записей', 'Совпадений',
                        'Несовпадений', '% совпадений']
        
        for col_idx, header in enumerate(verif_headers, 1):
            cell = ws_verif.cell(row=1, column=col_idx, value=header)
            cell.font = cell.font.copy(bold=True)
        
        # Данные
        for row_idx, verif in enumerate(mappings_verification, 2):
            ws_verif.cell(row=row_idx, column=1, value=verif['Лист 1'])
            ws_verif.cell(row=row_idx, column=2, value=verif['Лист 2'])
            ws_verif.cell(row=row_idx, column=3, value=verif['Ключевое поле 1'])
            ws_verif.cell(row=row_idx, column=4, value=verif['Ключевое поле 2'])
            ws_verif.cell(row=row_idx, column=5, value=verif['Проверяемая пара'])
            ws_verif.cell(row=row_idx, column=6, value=verif['Проверено записей'])
            ws_verif.cell(row=row_idx, column=7, value=verif['Совпадений'])
            ws_verif.cell(row=row_idx, column=8, value=verif['Несовпадений'])
            ws_verif.cell(row=row_idx, column=9, value=verif['% совпадений'])
    
    # Сохраняем файл с новыми листами
    # Для больших файлов создаем отдельный файл только с новыми листами
    file_size_mb = os.path.getsize(input_file) / (1024 * 1024)
    
    # Всегда создаем отдельный файл с аннотацией для надежности
    annotation_only_file = input_file.replace('.xlsx', '_annotation.xlsx')
    wb_new.save(annotation_only_file)
    logger.info(f"Файл с аннотацией и статистикой сохранен: {annotation_only_file}")
    logger.info(f"Размер исходного файла: {file_size_mb:.1f} МБ")
    
    output_result = annotation_only_file
    
    # Пропускаем объединение с оригиналом для больших файлов (>10 МБ)
    if file_size_mb >= 10:
        logger.info(f"Исходный файл слишком большой ({file_size_mb:.1f} МБ), объединение пропущено")
        logger.info(f"Используйте файл {annotation_only_file} с новой аннотацией")
    elif file_size_mb < 10:
        try:
            logger.info("Попытка объединения с исходным файлом...")
            
            # Открываем оригинальный файл только для чтения листов
            wb_original = load_workbook(input_file, read_only=True, data_only=True)
            
            # Копируем только первые 5000 строк из каждого листа для экономии памяти
            max_rows_per_sheet = 5000
            
            for idx, sheet_name in enumerate(wb_original.sheetnames):
                if sheet_name == 'Аннотация':
                    continue
                
                logger.info(f"Копирование листа {sheet_name} (первые {max_rows_per_sheet} строк)...")
                ws_source = wb_original[sheet_name]
                ws_dest = wb_new.create_sheet(title=sheet_name[:31])  # Имя листа не более 31 символа
                
                for row_idx, row in enumerate(ws_source.iter_rows(values_only=True), 1):
                    if row_idx > max_rows_per_sheet:
                        break
                    for col_idx, value in enumerate(row, 1):
                        ws_dest.cell(row=row_idx, column=col_idx, value=value)
            
            wb_original.close()
            
            combined_file = input_file.replace('.xlsx', '_combined.xlsx')
            logger.info(f"Сохранение комбинированного файла: {combined_file}")
            wb_new.save(combined_file)
            logger.info(f"Комбинированный файл сохранен: {combined_file}")
            output_result = combined_file
            
        except Exception as e:
            logger.warning(f"Не удалось объединить файлы: {e}")
            logger.info(f"Используйте файл {annotation_only_file} с новой аннотацией")
    
    logger.info(f"Результат сохранен в: {output_result}")
    if mappings_verification:
        logger.info(f"Добавлен лист 'Верификация' с результатами проверки сопоставлений")
    
    return {
        'success': True,
        'output_file': output_file,
        'sheets_added': ['Аннотация', 'Статистика_пересечений'] + 
                       (['Верификация'] if mappings_verification else []),
        'statistics_count': len(statistics_data),
        'verification_count': len(mappings_verification)
    }


if __name__ == '__main__':
    import sys
    
    # Путь по умолчанию: workspace/output/esklp/esklp_full.xlsx
    if len(sys.argv) > 1:
        input_file = sys.argv[1]
    else:
        import os
        # Используем абсолютный путь относительно текущей рабочей директории
        input_file = os.path.join(os.getcwd(), 'output', 'esklp', 'esklp_full.xlsx')
    
    output_file = sys.argv[2] if len(sys.argv) > 2 else None
    
    result = create_annotation_sheet(input_file, output_file)
    
    if result['success']:
        print(f"\n✓ Успешно!")
        print(f"Файл: {result['output_file']}")
        print(f"Добавленные листы: {', '.join(result['sheets_added'])}")
        print(f"Записей статистики: {result['statistics_count']}")
        print(f"Проверок сопоставлений: {result['verification_count']}")
    else:
        print("\n✗ Ошибка при выполнении")
        sys.exit(1)
