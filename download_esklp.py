#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Скрипт для загрузки полного справочника ЕСКЛП,
распаковки ZIP-архива и сохранения данных на лист 'ЕСКЛП' в Excel.

Особенности:
- Обход проблем с SSL-сертификатами госсайтов РФ
- Повторные попытки при ошибках сети
- Автоматический поиск и чтение всех файлов esklp_klp_* в архиве
- Сохранение строго на лист с именем "ЕСКЛП"
- Использование файла-шаблона esklp_full — шаблон.xlsx с готовым форматированием первых 5 строк
- Чтение данных начиная с 6-й строки (без заголовков и футера) из листа esklp_klp_*
- Удаление столбца "Номер регистрационного удостоверения" и пустых строк
- Подавление предупреждений openpyxl
- Подробное логирование в консоль и файл

Автор: Assistant
Дата: 2026
"""

# ============================================================================
# ИМПОРТЫ
# ============================================================================
import os
import io
import sys
import time
import logging
import tempfile
import zipfile
import warnings
import re
from pathlib import Path
from typing import Optional, List, Tuple
from urllib.parse import urlparse

# Отключаем предупреждения pandas о типах данных
warnings.filterwarnings('ignore', category=FutureWarning)

# Отключаем предупреждение openpyxl о default style
warnings.filterwarnings('ignore', message='Workbook contains no default style')

# Библиотеки для работы с сетью и данными
import requests
import pandas as pd
import urllib3
from openpyxl import load_workbook, Workbook

# Отключаем предупреждения о небезопасных HTTPS-запросах
# (необходимо для работы с госсайтами РФ)
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# ============================================================================
# НАСТРОЙКИ ЛОГИРОВАНИЯ
# ============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s | %(levelname)-8s | %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S',
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler('esklp_download.log', encoding='utf-8', mode='w')
    ]
)
logger = logging.getLogger(__name__)


# ============================================================================
# КОНФИГУРАЦИЯ
# ============================================================================
class Config:
    """Конфигурация скрипта"""

    # 🔗 Ссылка на скачивание ЕСКЛП (актуальная на 2026)
    ESKLP_URL = "https://esklp.egisz.rosminzdrav.ru/fs/public/download/3a250aaa-b686-40e3-9560-758ccce863aa"

    # 📁 Пути и имена файлов
    OUTPUT_DIR = Path("output/esklp")
    ZIP_FILENAME = "esklp_archive.zip"
    EXCEL_FILENAME = "esklp_full.xlsx"
    TEMPLATE_FILENAME = "esklp_full — шаблон.xlsx"  # Файл-шаблон с готовым форматированием
    TARGET_SHEET_NAME = "ЕСКЛП"  # Имя листа в итоговом Excel

    # ⏱️ Таймауты и повторные попытки
    DOWNLOAD_TIMEOUT = 120  # секунд на скачивание
    MAX_RETRIES = 3  # количество попыток при ошибке
    RETRY_DELAY = 5  # задержка между попытками (сек)

    # 🔐 Настройки SSL (для госсайтов)
    SSL_VERIFY = False  # False = отключить проверку сертификата
    SSL_CIPHERS = 'DEFAULT@SECLEVEL=1'  # Разрешить более старые шифры


# ============================================================================
# КЛАСС ЗАГРУЗЧИКА ЕСКЛП
# ============================================================================
class ESKLPDownloader:
    """Класс для загрузки, распаковки и сохранения справочника ЕСКЛП"""

    def __init__(self):
        """Инициализация: создаёт директорию, настраивает сессию"""
        Config.OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

        self.zip_path = Config.OUTPUT_DIR / Config.ZIP_FILENAME
        self.excel_path = Config.OUTPUT_DIR / Config.EXCEL_FILENAME
        self.template_path = Config.OUTPUT_DIR / Config.TEMPLATE_FILENAME

        # Настраиваем HTTP-сессию
        self.session = self._create_session()

        logger.info(f"📁 Рабочая директория: {Config.OUTPUT_DIR.resolve()}")

    def _create_session(self) -> requests.Session:
        """Создание настроенной HTTP-сессии"""
        session = requests.Session()

        # Заголовки, чтобы сервер не блокировал как бота
        session.headers.update({
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
            "Accept": "*/*",
            "Accept-Encoding": "gzip, deflate, br",
            "Connection": "keep-alive"
        })

        # Настройки SSL для работы с госсайтами РФ
        session.verify = Config.SSL_VERIFY

        if not Config.SSL_VERIFY:
            logger.debug("🔓 Проверка SSL-сертификата отключена")

        return session

    def download(self) -> bool:
        """
        Скачивание ZIP-архива с повторными попытками

        Returns:
            bool: True если скачивание успешно, иначе False
        """
        logger.info(f"🌐 Скачивание архива ЕСКЛП: {Config.ESKLP_URL}")

        for attempt in range(1, Config.MAX_RETRIES + 1):
            try:
                logger.debug(f"   Попытка #{attempt} из {Config.MAX_RETRIES}")

                response = self.session.get(
                    Config.ESKLP_URL,
                    stream=True,
                    timeout=Config.DOWNLOAD_TIMEOUT
                )
                response.raise_for_status()

                # Получаем размер файла для отображения прогресса
                total_size = int(response.headers.get('content-length', 0))
                downloaded = 0
                chunk_size = 8192  # 8 КБ chunks

                with open(self.zip_path, 'wb') as f:
                    for chunk in response.iter_content(chunk_size=chunk_size):
                        if chunk:
                            f.write(chunk)
                            downloaded += len(chunk)

                            # Логируем прогресс каждые 10 МБ
                            if downloaded % (10 * 1024 * 1024) == 0 and total_size > 0:
                                pct = downloaded / total_size * 100
                                logger.debug(
                                    f"   ⏳ {downloaded / 1024 ** 2:.1f} МБ / {total_size / 1024 ** 2:.1f} МБ ({pct:.1f}%)")

                # Проверка целостности
                if self.zip_path.stat().st_size == 0:
                    raise RuntimeError("Скачанный файл пустой")

                size_mb = self.zip_path.stat().st_size / 1024 ** 2
                logger.info(f"✅ Архив сохранён: {self.zip_path.name} ({size_mb:.2f} МБ)")
                return True

            except requests.exceptions.SSLError as e:
                logger.warning(f"⚠️ SSL-ошибка (попытка {attempt}): {type(e).__name__}")
                if attempt >= Config.MAX_RETRIES:
                    logger.error("❌ Не удалось подключиться: проблема с сертификатом")
                    return False

            except requests.exceptions.ConnectionError as e:
                logger.warning(f"⚠️ Ошибка соединения (попытка {attempt}): {e}")

            except requests.exceptions.Timeout:
                logger.warning(f"⚠️ Таймаут соединения (попытка {attempt})")

            except requests.exceptions.HTTPError as e:
                code = e.response.status_code if e.response else 'N/A'
                logger.error(f"❌ HTTP-ошибка {code}: {e}")
                return False  # Не повторяем при 403/404

            except Exception as e:
                logger.warning(f"⚠️ Неожиданная ошибка (попытка {attempt}): {type(e).__name__}: {e}")

            # Задержка перед следующей попыткой
            if attempt < Config.MAX_RETRIES:
                logger.info(f"   ⏳ Ожидание {Config.RETRY_DELAY} сек перед повтором...")
                time.sleep(Config.RETRY_DELAY)

        logger.error("❌ Не удалось скачать архив после всех попыток")
        return False

    def extract_and_load(self) -> pd.DataFrame:
        """
        Распаковка ZIP и чтение всех файлов esklp_klp_* в DataFrame

        Returns:
            pd.DataFrame или None: Объединённый DataFrame из всех файлов esklp_klp_*
        """
        if not self.zip_path.exists():
            logger.error("❌ ZIP-архив не найден. Сначала выполните download()")
            return None

        dataframes: List[pd.DataFrame] = []
        files_processed = 0

        with tempfile.TemporaryDirectory() as tmpdir:
            try:
                # Распаковка архива
                with zipfile.ZipFile(self.zip_path, 'r') as z:
                    z.extractall(tmpdir)
                    logger.info(f"📦 Архив распакован во временную директорию")

                # Рекурсивный поиск всех .xlsx файлов
                xlsx_files = list(Path(tmpdir).rglob("*.xlsx"))
                xlsx_files += list(Path(tmpdir).rglob("*.xls"))  # На всякий случай

                if not xlsx_files:
                    # Показываем что внутри архива для отладки
                    with zipfile.ZipFile(self.zip_path, 'r') as z:
                        contents = z.namelist()[:20]
                    logger.error(f"❌ Не найдено файлов .xlsx/.xls в архиве. Содержимое: {contents}")
                    return None

                logger.info(f"🔍 Найдено файлов для обработки: {len(xlsx_files)}")

                for file_path in xlsx_files:
                    logger.info(f"   📄 Обработка: {file_path.name}")
                    
                    # Обрабатываем ТОЛЬКО файлы esklp_klp_*
                    if not re.match(r'esklp_klp_.*\.xlsx', file_path.name, re.IGNORECASE):
                        logger.info(f"   ⏭️ Пропущен файл (не esklp_klp_*): {file_path.name}")
                        continue
                    
                    xls = None
                    try:
                        # Открываем книгу для получения списка листов
                        xls = pd.ExcelFile(file_path, engine='openpyxl')
                        
                        # Ищем лист с названием esklp_klp_*
                        target_sheet = None
                        for sheet_name in xls.sheet_names:
                            if re.match(r'esklp_klp_.*', sheet_name, re.IGNORECASE):
                                target_sheet = sheet_name
                                break
                        
                        if target_sheet is None:
                            logger.debug(f"      ⏭️ Пропущен файл (нет листа esklp_klp_*): {file_path.name}")
                            continue
                        
                        logger.debug(f"      → Лист: {target_sheet}")
                        
                        # Читаем данные начиная с 6-й строки (пропускаем первые 5 строк заголовка)
                        # header=None означает, что не будет автоматического определения заголовков
                        df = pd.read_excel(xls, sheet_name=target_sheet, engine='openpyxl', skiprows=5, header=None)
                        
                        if df.empty:
                            logger.debug(f"      ⚪ Пропущен пустой лист")
                            continue
                        
                        # Удаляем последние строки (футер) - строки где все значения NaN или пустые
                        # Ищем последнюю строку с данными и обрезаем всё после неё
                        last_valid_idx = df.last_valid_index()
                        if last_valid_idx is not None:
                            df = df.iloc[:last_valid_idx + 1]
                        
                        # Добавляем метаданные для отладки
                        df['_esklp_source_file'] = file_path.name
                        df['_esklp_source_sheet'] = target_sheet

                        dataframes.append(df)
                        files_processed += 1
                    except ValueError as e:
                        if "Excel file format cannot be determined" in str(e):
                            logger.warning(f"   ⚠️ Неверный формат файла: {file_path.name}")
                        else:
                            logger.warning(f"   ⚠️ Ошибка чтения листа: {e}")
                    except Exception as e:
                        logger.warning(f"   ⚠️ Ошибка обработки {file_path.name}: {type(e).__name__}: {e}")
                    finally:
                        # Явно закрываем файл, чтобы освободить дескриптор (важно для Windows)
                        if xls is not None:
                            xls.close()

            except zipfile.BadZipFile:
                logger.error("❌ Файл повреждён или не является ZIP-архивом")
                return None

            except Exception as e:
                logger.error(f"❌ Ошибка распаковки: {type(e).__name__}: {e}")
                return None

        if not dataframes:
            logger.warning("⚠️ Не удалось извлечь данные из архива")
            return None

        # Объединение всех найденных данных
        if len(dataframes) > 1:
            logger.info(f"🔗 Объединение {len(dataframes)} блоков данных...")
            combined_df = pd.concat(dataframes, ignore_index=True, sort=False)
        else:
            combined_df = dataframes[0]

        # Предварительная очистка
        combined_df.columns = combined_df.columns.astype(str).str.strip()
        combined_df = combined_df.dropna(how='all')  # Удаляем полностью пустые строки
        
        # Удаляем строки, где нет номера регистрационного удостоверения (пустые значения и пустые строки в этом столбце)
        if 'Номер регистрационного удостоверения' in combined_df.columns:
            # Заменяем пустые строки на NaN для корректного удаления
            combined_df['Номер регистрационного удостоверения'] = combined_df['Номер регистрационного удостоверения'].replace('', pd.NA)
            combined_df = combined_df.dropna(subset=['Номер регистрационного удостоверения'])
        
        # Удаляем столбец "Номер регистрационного удостоверения" если он существует
        combined_df = combined_df.drop(columns=['Номер регистрационного удостоверения'], errors='ignore')

        logger.info(f"📊 Результат: {len(combined_df):,} строк × {len(combined_df.columns)} колонок")
        return combined_df

    def save_to_excel(self, df: pd.DataFrame) -> bool:
        """
        Сохранение DataFrame на целевой лист Excel с использованием файла-шаблона.
        Первые 5 строк заголовка берутся из файла-шаблона esklp_full — шаблон.xlsx.
        Данные записываются начиная с 6-й строки.

        Args:
            df: DataFrame с данными ЕСКЛП (без заголовков, только данные)

        Returns:
            bool: True если сохранение успешно
        """
        try:
            logger.info(f"💾 Сохранение на лист '{Config.TARGET_SHEET_NAME}'...")

            # Создаём копию, чтобы не модифицировать исходный df
            df_export = df.copy()

            # Опционально: удалить служебные колонки перед сохранением
            df_export = df_export.drop(columns=['_esklp_source_file', '_esklp_source_sheet'], errors='ignore')

            # Проверяем наличие файла-шаблона
            if not self.template_path.exists():
                logger.error(f"❌ Файл-шаблон не найден: {self.template_path}")
                logger.error("💡 Убедитесь, что файл 'esklp_full — шаблон.xlsx' существует в директории {Config.OUTPUT_DIR}")
                return False

            # Загружаем файл-шаблон
            logger.info(f"📋 Использование шаблона: {self.template_path.name}")
            wb = load_workbook(self.template_path)
            
            # Получаем или создаём целевой лист
            if Config.TARGET_SHEET_NAME in wb.sheetnames:
                ws = wb[Config.TARGET_SHEET_NAME]
            else:
                ws = wb.active
                ws.title = Config.TARGET_SHEET_NAME
            
            # Очищаем всё, что есть после 5-й строки (если есть старые данные)
            for row_idx in range(6, ws.max_row + 1):
                for col_idx in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = None
            
            # Записываем данные ЕСКЛП начиная с 6-й строки (после заголовка из шаблона)
            data_start_row = 6
            
            for row_idx, row in enumerate(df_export.values, data_start_row):
                for col_idx, value in enumerate(row, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Закрепляем область после заголовка (после 5-й строки)
            ws.freeze_panes = 'A6'
            
            # Сохраняем книгу
            wb.save(self.excel_path)
            wb.close()

            size_mb = self.excel_path.stat().st_size / 1024 ** 2
            logger.info(f"✅ Файл сохранён: {self.excel_path} ({size_mb:.2f} МБ)")

            # Выводим превью колонок
            cols_preview = list(df_export.columns[:10])
            if len(df_export.columns) > 10:
                cols_preview.append(f"... и ещё {len(df_export.columns) - 10}")
            logger.info(f"📋 Колонки ЕСКЛП: {cols_preview}")
            
            # Показываем список листов
            result_wb = load_workbook(self.excel_path, read_only=True)
            logger.info(f"📑 Листы в файле: {result_wb.sheetnames}")
            result_wb.close()

            return True

        except PermissionError:
            logger.error(f"❌ Нет прав на запись в {self.excel_path}")
            logger.error("💡 Закройте файл в Excel и повторите попытку")
            return False
        except Exception as e:
            logger.error(f"❌ Ошибка сохранения Excel: {type(e).__name__}: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return False


    def run(self) -> bool:
        """
        Полный цикл загрузки: Скачать → Распаковать → Сохранить

        Returns:
            bool: True если все этапы выполнены успешно
        """
        logger.info("🚀 Запуск загрузки справочника ЕСКЛП")
        logger.info(f"🔗 URL: {Config.ESKLP_URL}")
        start_time = time.time()

        # Этап 1: Скачивание
        if not self.download():
            return False

        # Этап 2: Распаковка и чтение (возвращает DataFrame)
        df = self.extract_and_load()
        
        if df is None:
            logger.error("❌ Нет данных для сохранения")
            return False

        # Этап 3: Сохранение в Excel (с использованием шаблона)
        if not self.save_to_excel(df):
            return False

        # Итоговая статистика
        elapsed = time.time() - start_time
        logger.info(f"⏱️ Время выполнения: {elapsed:.1f} сек.")
        logger.info(f"🎉 Загрузка завершена успешно!")

        return True


    def cleanup(self):
        """Очистка: удаление временного ZIP-архива (опционально)"""
        if self.zip_path.exists():
            try:
                self.zip_path.unlink()
                logger.debug(f"🗑️ Удалён временный архив: {self.zip_path.name}")
            except Exception as e:
                logger.warning(f"⚠️ Не удалось удалить архив: {e}")


# ============================================================================
# ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ
# ============================================================================

def check_internet_connection(url: str = "https://google.com", timeout: int = 5) -> bool:
    """Проверка доступности интернета"""
    try:
        requests.get(url, timeout=timeout, verify=False)
        return True
    except:
        return False


def print_system_info():
    """Вывод информации о системе для отладки"""
    logger.debug("🔧 Информация о системе:")
    logger.debug(f"   Python: {sys.version.split()[0]}")
    logger.debug(f"   Platform: {sys.platform}")
    logger.debug(f"   Requests: {requests.__version__}")
    logger.debug(f"   Pandas: {pd.__version__}")
    logger.debug(f"   Working dir: {Path.cwd()}")


# ============================================================================
# ТОЧКА ВХОДА
# ============================================================================

if __name__ == "__main__":
    try:
        # Вывод информации о системе (для отладки)
        if '--debug' in sys.argv or '-v' in sys.argv:
            logging.getLogger().setLevel(logging.DEBUG)
            print_system_info()

        # Проверка подключения к интернету
        if not check_internet_connection():
            logger.error("❌ Нет подключения к интернету")
            sys.exit(1)

        # Запуск загрузчика
        downloader = ESKLPDownloader()
        success = downloader.run()

        # Очистка временных файлов (раскомментируйте если нужно)
        # downloader.cleanup()

        # Вывод результата
        if success:
            print(f"\n{'=' * 60}")
            print(f"🎉 ГОТОВО! Файл сохранён:")
            print(f"   {downloader.excel_path.resolve()}")
            print(f"{'=' * 60}")
            sys.exit(0)
        else:
            print(f"\n❌ Загрузка завершена с ошибками.")
            print(f"📄 Подробности в логе: esklp_download.log")
            sys.exit(1)

    except KeyboardInterrupt:
        logger.warning("\n⚠️ Прервано пользователем (Ctrl+C)")
        sys.exit(130)
    except Exception as e:
        logger.critical(f"💥 Критическая ошибка: {type(e).__name__}: {e}", exc_info=True)
        sys.exit(2)
